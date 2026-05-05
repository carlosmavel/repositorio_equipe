# utils.py

import bleach
import re
import unicodedata
from typing import Any, TypedDict

import os
import json
from docx import Document
import openpyxl
import xlrd
from odf import opendocument
from odf.text import P
import logging
import time
import warnings
from cryptography.utils import CryptographyDeprecationWarning

# Suprime avisos depreciação do ARC4 vindos do pypdf/cryptography durante o
# bootstrap da aplicação para evitar ruído no console do Flask.
warnings.filterwarnings(
    "ignore",
    category=CryptographyDeprecationWarning,
    message=r"ARC4 has been moved to cryptography\.hazmat\.decrepit\.ciphers\.algorithms\.ARC4.*",
)
try:
    import cv2
    import numpy as np
except Exception:  # pragma: no cover
    cv2 = None
    np = None
try:
    from pdf2image import convert_from_path
except Exception:  # pragma: no cover
    convert_from_path = None
try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None
try:
    import pytesseract
except Exception:  # pragma: no cover
    pytesseract = None
try:
    from PIL import Image
except Exception:  # pragma: no cover
    Image = None

try:
    from .database import db  # type: ignore  # pragma: no cover
    from .models import OrdemServico  # type: ignore  # pragma: no cover
except ImportError:  # pragma: no cover
    from core.database import db  # type: ignore
    from core.models import OrdemServico  # type: ignore

from sqlalchemy import select

logger = logging.getLogger(__name__)
_RESERVED_LOG_RECORD_FIELDS = set(logging.makeLogRecord({}).__dict__.keys())




def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value or "")
    return ''.join(ch for ch in normalized if unicodedata.category(ch) != 'Mn')


def build_like_pattern(term: str) -> str:
    cleaned = (term or '').strip()
    if not cleaned:
        return ''

    if '%' in cleaned:
        if not cleaned.startswith('%'):
            cleaned = f"%{cleaned}"
        if not cleaned.endswith('%'):
            cleaned = f"{cleaned}%"
        return cleaned

    return f"%{cleaned}%"

class ExtractedTextResult(TypedDict):
    text: str
    method: str
    total_pages: int
    pages_success: int
    pages_failed: int


def _build_extraction_result(
    text: str,
    *,
    method: str,
    total_pages: int = 1,
    pages_success: int | None = None,
    pages_failed: int | None = None,
) -> ExtractedTextResult:
    resolved_total = max(int(total_pages or 0), 1)
    resolved_success = pages_success if pages_success is not None else (1 if text.strip() else 0)
    resolved_success = max(min(int(resolved_success), resolved_total), 0)
    resolved_failed = pages_failed if pages_failed is not None else (resolved_total - resolved_success)
    resolved_failed = max(int(resolved_failed), 0)
    return {
        "text": text,
        "method": method,
        "total_pages": resolved_total,
        "pages_success": resolved_success,
        "pages_failed": resolved_failed,
    }


def build_article_log_context(
    *,
    user_id: int | None = None,
    route: str | None = None,
    action: str | None = None,
    article_id: int | None = None,
    attachment_id: int | None = None,
    boletim_id: int | None = None,
    filename: str | None = None,
    file_size: int | None = None,
    mime_type: str | None = None,
    ocr_status: str | None = None,
    attempt: int | None = None,
    progress_id: str | None = None,
    correlation_id: str | None = None,
    article_title: str | None = None,
    attachment_count: int | None = None,
    reason: str | None = None,
    ocr_method: str | None = None,
    ocr_engine: str | None = None,
    ocr_page_count: int | None = None,
    ocr_pages_success: int | None = None,
    ocr_pages_failed: int | None = None,
    ocr_char_count: int | None = None,
) -> dict[str, Any]:
    """Monta payload de contexto padronizado para logs de artigo/anexo/OCR."""
    return {
        "user_id": user_id,
        "route": route,
        "action": action,
        "article_id": article_id,
        "attachment_id": attachment_id,
        "boletim_id": boletim_id,
        "filename": filename,
        "file_size": file_size,
        "mime_type": mime_type,
        "ocr_status": ocr_status,
        "attempt": attempt,
        "progress_id": progress_id,
        "correlation_id": correlation_id,
        "article_title": article_title,
        "attachment_count": attachment_count,
        "reason": reason,
        "ocr_method": ocr_method,
        "ocr_engine": ocr_engine,
        "ocr_page_count": ocr_page_count,
        "ocr_pages_success": ocr_pages_success,
        "ocr_pages_failed": ocr_pages_failed,
        "ocr_char_count": ocr_char_count,
    }


def log_article_event(
    logger_obj: logging.Logger,
    event: str,
    *,
    level: int = logging.INFO,
    **context: Any,
) -> None:
    """Emite evento estruturado e consistente para observabilidade."""
    payload = build_article_log_context(**context)
    safe_extra = {"event_context": payload}
    safe_extra.update({k: v for k, v in payload.items() if k not in _RESERVED_LOG_RECORD_FIELDS})
    logger_obj.log(level, event, extra=safe_extra)


def log_article_exception(
    logger_obj: logging.Logger,
    event: str,
    **context: Any,
) -> None:
    payload = build_article_log_context(**context)
    safe_extra = {"event_context": payload}
    safe_extra.update({k: v for k, v in payload.items() if k not in _RESERVED_LOG_RECORD_FIELDS})
    logger_obj.exception(event, extra=safe_extra)
#-------------------------------------------------------------------------------------------
# Configura o campo de texto para se comportar corretamente quando recebe tags HTML
#-------------------------------------------------------------------------------------------
def sanitize_html(text: str) -> str:
    allowed_tags = [
        'p', 'br', 'strong', 'em', 'u', 's', 'sub', 'sup',
        'ul', 'ol', 'li', 'blockquote', 'code', 'pre', 'a'
    ]
    allowed_attrs = {
        '*': ['class'],
        'a': ['href', 'title'],
    }

    return bleach.clean(
        text,
        tags=allowed_tags,
        attributes=allowed_attrs,
        strip=True,
        protocols=['http', 'https', 'mailto']
    )

"""
Funções utilitárias do projeto.

Este módulo recebeu melhorias relacionadas ao pós-processamento do OCR. Toda a
extração existente permanece inalterada e os novos comportamentos são opcionais
e executados somente quando parâmetros específicos são informados.
"""

#-------------------------------------------------------------------------------------------
# Extração de texto dos anexos
#-------------------------------------------------------------------------------------------
def extract_text(path: str, **ocr_options) -> str | ExtractedTextResult:
    """Extrai texto de vários formatos de arquivo.

    Parâmetros adicionais são encaminhados para as rotinas de OCR quando o
    arquivo é um PDF ou imagem. O comportamento padrão permanece idêntico ao
    anterior quando nenhum parâmetro extra é fornecido.
    """
    progress_callback = ocr_options.pop("progress_callback", None)
    return_metadata = ocr_options.pop("return_metadata", False)

    ext = os.path.splitext(path)[1].lower()
    text_parts: list[str] = []

    # TXT
    if ext == '.txt':
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            result = _build_extraction_result(f.read(), method='plain_text')
            return result if return_metadata else result["text"]

    # DOCX
    if ext == '.docx':
        doc = Document(path)
        for para in doc.paragraphs:
            text_parts.append(para.text)
        result = _build_extraction_result('\n'.join(text_parts), method='document_text')
        return result if return_metadata else result["text"]

    # XLSX
    if ext == '.xlsx':
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        text_parts.append(str(cell))
        result = _build_extraction_result('\n'.join(text_parts), method='document_text')
        return result if return_metadata else result["text"]

    # XLS (Excel antigo)
    if ext == '.xls':
        wb = xlrd.open_workbook(path)
        for sheet in wb.sheets():
            for rx in range(sheet.nrows):
                for cx in range(sheet.ncols):
                    cell = sheet.cell(rx, cx).value
                    if cell:
                        text_parts.append(str(cell))
        result = _build_extraction_result('\n'.join(text_parts), method='document_text')
        return result if return_metadata else result["text"]

    # ODS (LibreOffice Calc)
    if ext == '.ods':
        doc = opendocument.load(path)
        for elem in doc.getElementsByType(P):
            text_parts.append(str(elem))
        result = _build_extraction_result('\n'.join(text_parts), method='document_text')
        return result if return_metadata else result["text"]

    # PDF
    if ext == '.pdf':
        return extract_text_from_pdf(
            path,
            progress_callback=progress_callback,
            return_metadata=return_metadata,
            **ocr_options,
        )

    # outros formatos não suportados
    result = _build_extraction_result('', method='unsupported_format')
    return result if return_metadata else result["text"]


def preprocess_image(img, debug_dir=None, page_idx=0):
    """Aplica etapas de pre-processamento utilizando OpenCV."""
    if not (cv2 and np):  # pragma: no cover - dependencias ausentes
        return img

    img_array = np.array(img)
    gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
    if debug_dir:
        cv2.imwrite(os.path.join(debug_dir, f"page_{page_idx}_gray.png"), gray)

    sharp = cv2.addWeighted(gray, 1.5, cv2.GaussianBlur(gray, (0, 0), 1.0), -0.5, 0)
    if debug_dir:
        cv2.imwrite(os.path.join(debug_dir, f"page_{page_idx}_sharp.png"), sharp)

    _, binary = cv2.threshold(sharp, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    if debug_dir:
        cv2.imwrite(os.path.join(debug_dir, f"page_{page_idx}_binary.png"), binary)

    return Image.fromarray(binary)


def _reconstruct_text(data: dict, width: int) -> str:
    """Reordena blocos de texto obtidos via ``image_to_data``."""
    items = []
    for i, txt in enumerate(data.get("text", [])):
        if not txt:
            continue
        items.append(
            (
                data["page_num"][i],
                data["top"][i],
                data["left"][i],
                txt,
            )
        )

    from itertools import groupby

    pages: list[str] = []
    for page, group in groupby(items, key=lambda x: x[0]):
        group_items = list(group)
        xs = [g[2] for g in group_items]
        column_split = None
        if xs:
            mid = width / 2
            left_count = sum(1 for x in xs if x < mid)
            right_count = sum(1 for x in xs if x >= mid)
            if left_count > 0 and right_count > 0 and left_count / len(xs) > 0.2 and right_count / len(xs) > 0.2:
                column_split = mid

        if column_split:
            left_items = sorted([g for g in group_items if g[2] < column_split], key=lambda x: (x[1], x[2]))
            right_items = sorted([g for g in group_items if g[2] >= column_split], key=lambda x: (x[1], x[2]))
            ordered = left_items + right_items
        else:
            ordered = sorted(group_items, key=lambda x: (x[1], x[2]))

        current_y = None
        page_lines: list[str] = []
        for _, top, left, text in ordered:
            if current_y is None or abs(top - current_y) > 10:
                page_lines.append(text)
                current_y = top
            else:
                page_lines[-1] += " " + text
        pages.append("\n".join(page_lines).strip())

    return "\n".join(pages)


def _clean_text(text: str) -> str:
    """Remove ruídos comuns do OCR e normaliza o texto."""
    text = re.sub(r"(?<!\S)[ºª|—](?!\S)", "", text)
    text = re.sub(r"\s[?<>]\s", " ", text)
    text = re.sub(r"(\w+)-\n(\w+)", r"\1\2\n", text)
    text = re.sub(r"[ ]{2,}", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def _detect_sparse(data: dict, width: int, height: int) -> bool:
    words = [t for t in data.get("text", []) if t.strip()]
    if not words:
        return True
    areas = [data["width"][i] * data["height"][i] for i, t in enumerate(data["text"]) if t.strip()]
    density = sum(areas) / float(width * height)
    return density < 0.02 or len(words) < 20


def extract_text_from_image(
    image,
    lang: str = "por",
    *,
    reorder: bool = False,
    clean: bool = False,
    detect_sparse: bool = False,
) -> str:
    """Realiza OCR na imagem usando pytesseract com pós-processamento opcional."""
    if not pytesseract:  # pragma: no cover - dependencias ausentes
        return ""

    if reorder or detect_sparse:
        data = pytesseract.image_to_data(
            image, lang=lang, config="--oem 3 --psm 6", output_type=pytesseract.Output.DICT
        )
        text = _reconstruct_text(data, image.width)
        if detect_sparse and _detect_sparse(data, image.width, image.height):
            aux = pytesseract.image_to_string(image, lang=lang, config="--oem 3 --psm 11")
            if aux:
                text = text + "\n" + aux.replace("\x0c", "").strip()
    else:
        text = pytesseract.image_to_string(
            image, lang=lang, config="--oem 3 --psm 6"
        )

    text = text.replace("\x0c", "").strip()
    if clean:
        text = _clean_text(text)
    return text


def extract_text_from_pdf(
    path: str,
    *,
    lang: str = "por",
    reorder: bool = False,
    clean: bool = False,
    detect_sparse: bool = False,
    progress_callback=None,
    return_metadata: bool = False,
) -> str | ExtractedTextResult:
    """Extrai texto de PDFs pesquisáveis ou via ``pdf2image`` + ``pytesseract``.

    Primeiro tenta recuperar o conteúdo diretamente (PDF pesquisável) para
    evitar OCR desnecessário. O reordenamento, a limpeza de texto e a detecção
    de layout esparso são opcionais e aplicados apenas quando os parâmetros
    correspondentes são habilitados. Com todos os parâmetros em ``False`` o
    comportamento é idêntico ao anterior.
    """
    started_at = time.monotonic()
    total_pages = _get_pdf_total_pages(path)
    logger.info("pdf_page_count_detected", extra={"ocr_page_count": total_pages})

    direct_text = _extract_pdf_text_without_ocr(path)
    direct_char_count = len(re.sub(r"\s+", "", direct_text or ""))
    direct_text_min_chars = 100
    min_chars_per_page = 50
    should_fallback = (
        not direct_text
        or direct_char_count < direct_text_min_chars
        or (total_pages > 0 and (direct_char_count / total_pages) < min_chars_per_page)
    )
    fallback_reason = None
    if not direct_text:
        fallback_reason = "direct_text_empty"
    elif direct_char_count < direct_text_min_chars:
        fallback_reason = "direct_text_below_min_chars"
    elif total_pages > 0 and (direct_char_count / total_pages) < min_chars_per_page:
        fallback_reason = "direct_text_below_min_chars_per_page"

    logger.info(
        "pdf_direct_text_evaluation",
        extra={
            "ocr_page_count": total_pages,
            "ocr_char_count": direct_char_count,
            "ocr_fallback": should_fallback,
            "ocr_fallback_reason": fallback_reason,
        },
    )

    if not should_fallback:
        result = _build_extraction_result(
            direct_text,
            method="pdf_direct_text",
            total_pages=total_pages,
            pages_success=total_pages if direct_text.strip() else 0,
            pages_failed=0 if direct_text.strip() else total_pages,
        )
        final_char_count = direct_char_count
        logger.info(
            "ocr_result_selected",
            extra={
                "ocr_method": result["method"],
                "ocr_engine": result["method"],
                "ocr_page_count": result["total_pages"],
                "ocr_pages_success": result["pages_success"],
                "ocr_pages_failed": result["pages_failed"],
                "ocr_char_count": final_char_count,
                "ocr_status": "concluido",
            },
        )
        logger.info(
            "ocr_processing_finished",
            extra={
                "ocr_method": result["method"],
                "ocr_engine": result["method"],
                "ocr_processing_time_seconds": round(time.monotonic() - started_at, 4),
                "ocr_status": "concluido",
            },
        )
        return result if return_metadata else result["text"]

    if not (convert_from_path and Image and pytesseract):
        logger.warning("pdf2image, PIL ou pytesseract indisponivel para %s", path)
        result = _build_extraction_result("", method="pdf_ocr_page_by_page")
        return result if return_metadata else result["text"]
    logger.info(
        "ocr_fallback_started",
        extra={
            "ocr_page_count": total_pages,
            "ocr_fallback_reason": fallback_reason or "direct_text_low_yield",
            "ocr_status": "baixo_aproveitamento",
        },
    )
    try:
        images = convert_from_path(path, dpi=300)
    except Exception as e:  # pragma: no cover - erro ao converter
        logger.error("Erro ao converter PDF %s: %s", path, e)
        result = _build_extraction_result("", method="pdf_ocr_page_by_page")
        return result if return_metadata else result["text"]
    text_parts: list[str] = []
    logger.info(
        "pdf_ocr_page_by_page",
        extra={
            "ocr_method": "pdf_ocr_page_by_page",
            "ocr_engine": "pdf_ocr_page_by_page",
            "ocr_page_count": total_pages,
            "ocr_status": "processando",
        },
    )
    for i, img in enumerate(images, start=1):
        try:
            pre = preprocess_image(img, page_idx=i)
            text_parts.append(
                extract_text_from_image(
                    pre,
                    lang=lang,
                    reorder=reorder,
                    clean=clean,
                    detect_sparse=detect_sparse,
                )
            )
            logger.info("Pagina %s processada com sucesso", i)
            if progress_callback:
                percent = (i / total_pages) * 100 if total_pages else None
                progress_callback({
                    "message": f"Página {i} processada com sucesso",
                    "percent": percent,
                })
        except Exception as e:  # pragma: no cover
            logger.error("Erro no OCR da pagina %s do PDF %s: %s", i, path, e)
            text_parts.append("")
    pages_success = sum(1 for txt in text_parts if (txt or "").strip())
    final_text = "\n".join(text_parts)
    final_char_count = len(re.sub(r"\s+", "", final_text or ""))
    pages_failed = max(total_pages - pages_success, 0)
    logger.info(
        "ocr_result_selected",
        extra={
            "ocr_method": "pdf_ocr_page_by_page",
            "ocr_engine": "pdf_ocr_page_by_page",
            "ocr_page_count": total_pages,
            "ocr_pages_success": pages_success,
            "ocr_pages_failed": pages_failed,
            "ocr_char_count": final_char_count,
            "ocr_status": "concluido" if final_char_count else "baixo_aproveitamento",
        },
    )
    logger.info(
        "ocr_processing_finished",
        extra={
            "ocr_method": "pdf_ocr_page_by_page",
            "ocr_engine": "pdf_ocr_page_by_page",
            "ocr_processing_time_seconds": round(time.monotonic() - started_at, 4),
            "ocr_status": "concluido" if final_char_count else "baixo_aproveitamento",
        },
    )
    result = _build_extraction_result(
        final_text,
        method="pdf_ocr_page_by_page",
        total_pages=total_pages,
        pages_success=pages_success,
        pages_failed=pages_failed,
    )
    return result if return_metadata else result["text"]


def _get_pdf_total_pages(path: str) -> int:
    if not PdfReader:  # pragma: no cover
        return 1
    try:
        return max(len(PdfReader(path).pages), 1)
    except Exception as e:  # pragma: no cover
        logger.warning("Falha ao detectar total de paginas de %s: %s", path, e)
        return 1


def _extract_pdf_text_without_ocr(path: str, sample_pages: int = 3) -> str:
    """Tenta extrair texto de PDFs pesquisáveis antes de acionar o OCR.

    O objetivo é evitar o processamento custoso de OCR em arquivos que já
    contêm texto acessível. Caso nenhuma página amostrada possua texto ou
    ocorra qualquer erro, retorna uma string vazia para que o fluxo de OCR
    seja acionado normalmente.
    """
    if not PdfReader:  # pragma: no cover - dependencias ausentes
        return ""

    def _has_meaningful_text(text: str) -> bool:
        return len(re.sub(r"\s+", "", text)) >= 40

    try:
        reader = PdfReader(path)
    except Exception as e:  # pragma: no cover - falha ao abrir
        logger.warning("Falha ao abrir PDF %s para leitura direta: %s", path, e)
        return ""

    has_text = False
    for idx, page in enumerate(reader.pages):
        try:
            snippet = (page.extract_text() or "").strip()
        except Exception as e:  # pragma: no cover - erro específico de página
            logger.info("Erro ao ler pagina %s de %s sem OCR: %s", idx + 1, path, e)
            snippet = ""
        if _has_meaningful_text(snippet):
            has_text = True
            break
        if idx + 1 >= sample_pages:
            break

    if not has_text:
        return ""

    pages_text: list[str] = []
    for idx, page in enumerate(reader.pages):
        try:
            page_text = page.extract_text() or ""
        except Exception as e:  # pragma: no cover
            logger.info("Erro ao extrair texto da pagina %s de %s: %s", idx + 1, path, e)
            page_text = ""
        pages_text.append(page_text.strip())
    joined_text = "\n".join(pages_text).strip()
    if not _has_meaningful_text(joined_text):
        return ""

    return "\n".join(text for text in pages_text if text)

import secrets
import string

DEFAULT_NEW_USER_PASSWORD = 'Mudanca123!'


def gerar_codigo_os() -> str:
    """Gera um código sequencial para Ordem de Serviço.

    O código possui um prefixo alfabético seguido de seis dígitos. Exemplo:
    ``A000001``. Quando o número atinge ``999999`` o prefixo é avançado para a
    próxima letra. A consulta utiliza ``SELECT ... FOR UPDATE`` para evitar
    condições de corrida; portanto, é esperado que esta função seja chamada
    dentro de uma transação ativa.
    """

    stmt = (
        select(OrdemServico.codigo)
        .order_by(OrdemServico.codigo.desc())
        .limit(1)
        .with_for_update()
    )
    ultimo_codigo = db.session.execute(stmt).scalar()

    if ultimo_codigo:
        prefixo = ultimo_codigo[0]
        numero = int(ultimo_codigo[1:]) + 1
        if numero > 999999:
            prefixo = chr(ord(prefixo) + 1)
            numero = 1
    else:
        prefixo, numero = 'A', 1

    if ord(prefixo) > ord('Z'):
        raise ValueError('Limite de códigos atingido')

    codigo = f"{prefixo}{numero:06d}"

    # Garantia adicional de unicidade
    while db.session.execute(
        select(OrdemServico.id).filter_by(codigo=codigo)
    ).scalar():
        numero += 1
        if numero > 999999:
            prefixo = chr(ord(prefixo) + 1)
            numero = 1
            if ord(prefixo) > ord('Z'):
                raise ValueError('Limite de códigos atingido')
        codigo = f"{prefixo}{numero:06d}"

    return codigo


def password_meets_requirements(password: str) -> bool:
    return (
        len(password) >= 8
        and re.search(r"[A-Z]", password)
        and re.search(r"[a-z]", password)
        and re.search(r"[0-9]", password)
        and re.search(r"[!@#$%^&*(),.?\":{}|<>]", password)
    )

def generate_random_password(length=12):
    """Gera uma senha aleatória com letras, números e caracteres especiais."""
    alphabet = string.ascii_letters + string.digits + '!@#$%^&*()-_'
    return ''.join(secrets.choice(alphabet) for _ in range(length))

from itsdangerous import URLSafeTimedSerializer
from flask import current_app
from .email_service import send_email

try:
    from .enums import Permissao  # type: ignore  # pragma: no cover
except ImportError:  # pragma: no cover - fallback for direct execution
    from core.enums import Permissao

def _password_token_serializer():
    secret = current_app.config.get("PASSWORD_RESET_SECRET") or current_app.secret_key
    return URLSafeTimedSerializer(secret)


def generate_token(user_id: int, action: str, expires_sec: int = 3600) -> str:
    """Gera um token seguro para ações como reset ou criação de senha."""
    serializer = _password_token_serializer()
    return serializer.dumps({'user_id': user_id, 'action': action})

def confirm_token(token: str, expiration: int = 3600):
    """Valida e decodifica o token, retornando o payload ou None."""
    serializer = _password_token_serializer()
    try:
        return serializer.loads(token, max_age=expiration)
    except Exception:
        return None

def user_can_edit_article(user, article):
    """Verifica se o usuário pode editar determinado artigo."""
    try:
        from .models import Article  # type: ignore  # pragma: no cover
    except ImportError:  # pragma: no cover - fallback for direct execution
        from core.models import Article

    if user is None:
        return False

    if not isinstance(article, Article):
        return False

    if user.has_permissao("admin") or user.id == article.user_id:
        return True

    if user.has_permissao(Permissao.ARTIGO_EDITAR_TODAS.value):
        return True

    if user.has_permissao(Permissao.ARTIGO_EDITAR_INSTITUICAO.value):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.instituicao_id == user_est.instituicao_id:
            return True

    if user.has_permissao(Permissao.ARTIGO_EDITAR_ESTABELECIMENTO.value):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.estabelecimento_id == user_est.id:
            return True

    if user.has_permissao(Permissao.ARTIGO_EDITAR_SETOR.value):
        if article.setor_id == user.setor_id:
            return True
        if user.extra_setores.filter_by(id=article.setor_id).count():
            return True

    if user.has_permissao(Permissao.ARTIGO_EDITAR_CELULA.value):
        if article.celula_id == user.celula_id:
            return True
        if user.extra_celulas.filter_by(id=article.celula_id).count():
            return True

    return False


def user_can_approve_article(user, article):
    """Verifica se o usuário pode aprovar determinado artigo."""
    try:
        from .models import Article  # type: ignore  # pragma: no cover
    except ImportError:  # pragma: no cover - fallback for direct execution
        from core.models import Article

    if user is None:
        return False

    if not isinstance(article, Article):
        return False

    # Nunca permitir aprovação do próprio artigo
    if user.id == getattr(article, "user_id", None):
        return False

    if user.has_permissao("admin") or user.has_permissao(Permissao.ARTIGO_APROVAR_TODAS.value):
        return True

    if user.has_permissao(Permissao.ARTIGO_APROVAR_INSTITUICAO.value):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.instituicao_id == user_est.instituicao_id:
            return True

    if user.has_permissao(Permissao.ARTIGO_APROVAR_ESTABELECIMENTO.value):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.estabelecimento_id == user_est.id:
            return True

    if user.has_permissao(Permissao.ARTIGO_APROVAR_SETOR.value):
        if article.setor_id == user.setor_id:
            return True
        if user.extra_setores.filter_by(id=article.setor_id).count():
            return True

    if user.has_permissao(Permissao.ARTIGO_APROVAR_CELULA.value):
        if article.celula_id == user.celula_id:
            return True
        if user.extra_celulas.filter_by(id=article.celula_id).count():
            return True

    return False


def user_can_review_article(user, article):
    """Verifica se o usuário pode revisar determinado artigo."""
    try:
        from .models import Article  # type: ignore  # pragma: no cover
    except ImportError:  # pragma: no cover - fallback for direct execution
        from core.models import Article

    if user is None:
        return False

    if not isinstance(article, Article):
        return False

    if user.has_permissao("admin") or user.has_permissao(Permissao.ARTIGO_REVISAR_TODAS.value):
        return True

    if user.has_permissao(Permissao.ARTIGO_REVISAR_INSTITUICAO.value):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.instituicao_id == user_est.instituicao_id:
            return True

    if user.has_permissao(Permissao.ARTIGO_REVISAR_ESTABELECIMENTO.value):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.estabelecimento_id == user_est.id:
            return True

    if user.has_permissao(Permissao.ARTIGO_REVISAR_SETOR.value):
        if article.setor_id == user.setor_id:
            return True
        if user.extra_setores.filter_by(id=article.setor_id).count():
            return True

    if user.has_permissao(Permissao.ARTIGO_REVISAR_CELULA.value):
        if article.celula_id == user.celula_id:
            return True
        if user.extra_celulas.filter_by(id=article.celula_id).count():
            return True

    return False

def user_can_view_article(user, article):
    """Verifica se o usuário tem permissão para visualizar o artigo."""
    try:
        from .models import Article  # type: ignore  # pragma: no cover
    except ImportError:  # pragma: no cover - fallback for direct execution
        from core.models import Article
    try:
        from .enums import ArticleVisibility
    except ImportError:  # pragma: no cover - fallback for direct execution
        from core.enums import ArticleVisibility

    if user is None:
        return False

    if not isinstance(article, Article):
        return False

    if user_can_edit_article(user, article):
        return True

    if user.has_permissao("admin"):
        return True

    if (
        user.has_permissao(Permissao.ARTIGO_APROVAR_TODAS.value)
        or user.has_permissao(Permissao.ARTIGO_REVISAR_TODAS.value)
    ):
        return True

    if (
        user.has_permissao(Permissao.ARTIGO_APROVAR_INSTITUICAO.value)
        or user.has_permissao(Permissao.ARTIGO_REVISAR_INSTITUICAO.value)
    ):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.instituicao_id == user_est.instituicao_id:
            return True

    if (
        user.has_permissao(Permissao.ARTIGO_APROVAR_ESTABELECIMENTO.value)
        or user.has_permissao(Permissao.ARTIGO_REVISAR_ESTABELECIMENTO.value)
    ):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.estabelecimento_id == user_est.id:
            return True

    if (
        user.has_permissao(Permissao.ARTIGO_APROVAR_SETOR.value)
        or user.has_permissao(Permissao.ARTIGO_REVISAR_SETOR.value)
    ):
        if article.setor_id == user.setor_id or user.extra_setores.filter_by(id=article.setor_id).count():
            return True

    if (
        user.has_permissao(Permissao.ARTIGO_APROVAR_CELULA.value)
        or user.has_permissao(Permissao.ARTIGO_REVISAR_CELULA.value)
    ):
        if article.celula_id == user.celula_id or user.extra_celulas.filter_by(id=article.celula_id).count():
            return True

    # Extras concedidos especificamente
    if article.extra_users.filter_by(id=user.id).count():
        return True
    if user.celula_id and article.extra_celulas.filter_by(id=user.celula_id).count():
        return True
    if any(article.extra_celulas.filter_by(id=c.id).count() for c in user.extra_celulas.all()):
        return True

    vis = article.visibility

    user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
    user_setor = user.setor or (user.celula.setor if user.celula else None)

    if vis is ArticleVisibility.INSTITUICAO:
        if user_est and article.instituicao_id == user_est.instituicao_id:
            return True
    elif vis is ArticleVisibility.ESTABELECIMENTO:
        if user_est and article.estabelecimento_id == user_est.id:
            return True
    elif vis is ArticleVisibility.SETOR:
        if user_setor and article.setor_id == user_setor.id:
            return True
        if article.setor_id and user.extra_setores.filter_by(id=article.setor_id).count():
            return True
    elif vis is ArticleVisibility.CELULA:
        # Usa a célula explícita de visibilidade se definida; caso contrário,
        # faz fallback para a célula do autor (para artigos antigos que não
        # possuam o campo vis_celula_id preenchido)
        cel_id = article.vis_celula_id or article.celula_id
        if user.celula_id == cel_id:
            return True
        # Considera usuários atribuídos a várias células
        if cel_id and user.extra_celulas.filter_by(id=cel_id).count():
            return True

    return False


def eligible_review_notification_users(article):
    """Retorna os usuários que devem ser notificados sobre a revisão do artigo."""
    try:
        from .models import User  # type: ignore  # pragma: no cover
    except ImportError:  # pragma: no cover - fallback for direct execution
        from core.models import User

    return [
        u for u in User.query.all()
        if u.id != getattr(article, "user_id", None)
        and user_can_approve_article(u, article)
    ]


def user_can_access_form_builder(user):
    """Verifica se o usuário tem acesso ao criador de formulários."""
    return bool(
        user
        and (
            getattr(user, 'pode_atender_os', False)
            or getattr(user, 'has_permissao', lambda _p: False)('admin')
        )
    )


def validar_fluxo_ramificacoes(estrutura):
    """Valida ramificações de formulários para evitar ciclos e destinos inválidos.

    Estrutura pode ser uma string JSON ou uma lista de blocos. Retorna
    ``(True, '')`` se estiver tudo correto ou ``(False, mensagem)`` em caso de
    conflito.
    """
    if isinstance(estrutura, str):
        try:
            data = json.loads(estrutura)
        except ValueError:
            return False, "Estrutura do formulário inválida."
    else:
        data = estrutura

    perguntas = []

    def coletar(itens):
        for item in itens:
            if item.get('tipo') == 'section':
                coletar(item.get('campos', []))
            else:
                perguntas.append(item)

    coletar(data)
    id_para_indice = {str(p['id']): idx for idx, p in enumerate(perguntas)}

    for idx, pergunta in enumerate(perguntas):
        for regra in pergunta.get('ramificacoes') or []:
            destino = regra.get('destino')
            if not destino or destino in ('next', 'end'):
                continue
            if destino not in id_para_indice:
                return False, (
                    f"Não é possível criar uma ramificação para a pergunta {destino}, "
                    "pois ela está inativa ou não existe."
                )
            dest_idx = id_para_indice[destino]
            if dest_idx <= idx:
                return False, (
                    f"Não é possível criar uma ramificação que retorna à Pergunta {dest_idx + 1}, "
                    "pois isso geraria um ciclo no formulário."
                )

    return True, ''
