# utils.py

import bleach
from bleach.css_sanitizer import CSSSanitizer
import re

import os
from docx import Document
import openpyxl
import xlrd
from odf import opendocument
from odf.text import P
import logging
try:
    from pdf2image import convert_from_path
except Exception:  # pragma: no cover
    convert_from_path = None
try:
    from paddleocr import PaddleOCR
except Exception:  # pragma: no cover
    PaddleOCR = None
try:
    import numpy as np
except Exception:  # pragma: no cover
    np = None
try:
    from PIL import Image
except Exception:  # pragma: no cover
    Image = None

logger = logging.getLogger(__name__)



OCR_ENGINE = None

def get_ocr_engine():
    """Retorna instancia unica do PaddleOCR configurada."""
    global OCR_ENGINE
    if OCR_ENGINE is None and PaddleOCR:
        OCR_ENGINE = PaddleOCR(use_angle_cls=True, lang="pt", use_gpu=False)
    return OCR_ENGINE


#-------------------------------------------------------------------------------------------
# Configura o campo de texto para se comportar corretamente quando recebe tags HTML
#-------------------------------------------------------------------------------------------
def sanitize_html(text: str) -> str:
    allowed_tags = [
        'p', 'br', 'strong', 'em', 'u', 's', 'sub', 'sup',
        'ul', 'ol', 'li', 'blockquote', 'code', 'pre',
        'a', 'img', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6',
        'table', 'thead', 'tbody', 'tr', 'th', 'td'
    ]
    allowed_attrs = {
        '*': ['class', 'style'],
        'a': ['href', 'title', 'target'],
        'img': ['src', 'alt', 'title', 'width', 'height'],
        'td': ['colspan', 'rowspan'],
        'th': ['colspan', 'rowspan']
    }
    # Configura quais propriedades CSS permitimos
    css_sanitizer = CSSSanitizer(
        allowed_css_properties=[
            'color',
            'background-color',
            'text-align',
            'width',
            'height',
            'border'
        ]
    )

    return bleach.clean(
        text,
        tags=allowed_tags,
        attributes=allowed_attrs,
        strip=True,
        css_sanitizer=css_sanitizer
    )

#-------------------------------------------------------------------------------------------
# Extração de texto dos anexos
#-------------------------------------------------------------------------------------------
def extract_text(path: str) -> str:
    """
    Extrai texto de vários formatos de arquivo:
    - .txt, .docx, .xlsx, .xls, .ods, .pdf
    Retorna todo o texto concatenado como string.
    """
    ext = os.path.splitext(path)[1].lower()
    text_parts = []

    # TXT
    if ext == '.txt':
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()

    # DOCX
    if ext == '.docx':
        doc = Document(path)
        for para in doc.paragraphs:
            text_parts.append(para.text)
        return '\n'.join(text_parts)

    # XLSX
    if ext == '.xlsx':
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        text_parts.append(str(cell))
        return '\n'.join(text_parts)

    # XLS (Excel antigo)
    if ext == '.xls':
        wb = xlrd.open_workbook(path)
        for sheet in wb.sheets():
            for rx in range(sheet.nrows):
                for cx in range(sheet.ncols):
                    cell = sheet.cell(rx, cx).value
                    if cell:
                        text_parts.append(str(cell))
        return '\n'.join(text_parts)

    # ODS (LibreOffice Calc)
    if ext == '.ods':
        doc = opendocument.load(path)
        for elem in doc.getElementsByType(P):
            text_parts.append(str(elem))
        return '\n'.join(text_parts)

    # PDF (texto ou imagem)
    if ext == '.pdf':
        return extract_text_from_pdf(path)

    # outros formatos não suportados
    return ''


def extract_text_from_pdf(path: str) -> str:
    """Extrai texto de PDFs usando pdf2image + PaddleOCR."""
    text_parts = []

    if not (convert_from_path and Image and np):
        logger.warning("pdf2image ou PIL numpy indisponivel para %s", path)
        return ""

    ocr_engine = get_ocr_engine()
    if not ocr_engine:
        logger.warning("PaddleOCR nao disponivel para %s", path)
        return ""

    try:
        images = convert_from_path(path, dpi=200)
    except Exception as e:  # pragma: no cover - erro ao converter
        logger.error("Erro ao converter PDF %s: %s", path, e)
        return ""

    for img in images:
        try:
            arr = np.array(img)
            result = ocr_engine.ocr(arr, cls=True)
            for res in result:
                if res and len(res) > 1:
                    text_parts.append(res[1][0])
        except Exception as e:  # pragma: no cover
            logger.error("Erro no OCR da pagina do PDF %s: %s", path, e)

    return "\n".join(text_parts)

import secrets
import string

DEFAULT_NEW_USER_PASSWORD = 'Mudanca123!'


def password_meets_requirements(password: str) -> bool:
    return (
        len(password) >= 8
        and re.search(r"[A-Z]", password)
        and re.search(r"[a-z]", password)
        and re.search(r"[0-9]", password)
        and re.search(r"[!@#$%^&*(),.?\":{}|<>]", password)
    )

def generate_random_password(length=12):
    """Gera uma senha aleatória com letras, números e caracteres especiais."""
    alphabet = string.ascii_letters + string.digits + '!@#$%^&*()-_'
    return ''.join(secrets.choice(alphabet) for _ in range(length))

from itsdangerous import URLSafeTimedSerializer
from flask import current_app
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail
import os

try:
    from .enums import Permissao  # type: ignore  # pragma: no cover
except ImportError:  # pragma: no cover - fallback for direct execution
    from enums import Permissao

def generate_token(user_id: int, action: str, expires_sec: int = 3600) -> str:
    """Gera um token seguro para ações como reset ou criação de senha."""
    serializer = URLSafeTimedSerializer(current_app.secret_key)
    return serializer.dumps({'user_id': user_id, 'action': action})

def confirm_token(token: str, expiration: int = 3600):
    """Valida e decodifica o token, retornando o payload ou None."""
    serializer = URLSafeTimedSerializer(current_app.secret_key)
    try:
        return serializer.loads(token, max_age=expiration)
    except Exception:
        return None

def send_email(to_email: str, subject: str, html_content: str) -> None:
    """Envia um e-mail utilizando o serviço SendGrid se configurado."""
    api_key = os.environ.get('SENDGRID_API_KEY')
    from_email = os.environ.get('EMAIL_FROM', 'no-reply@example.com')
    if not api_key:
        current_app.logger.warning('SendGrid API key não configurada.')
        return
    try:
        sg = SendGridAPIClient(api_key)
        message = Mail(from_email=from_email, to_emails=to_email,
                        subject=subject, html_content=html_content)
        sg.send(message)
    except Exception as e:
        current_app.logger.error(f'Erro ao enviar e-mail: {e}')


def user_can_edit_article(user, article):
    """Verifica se o usuário pode editar determinado artigo."""
    try:
        from .models import Article  # type: ignore  # pragma: no cover
    except ImportError:  # pragma: no cover - fallback for direct execution
        from models import Article

    if not isinstance(article, Article):
        return False

    if user.has_permissao("admin") or user.id == article.user_id:
        return True

    if user.has_permissao(Permissao.ARTIGO_EDITAR_TODAS.value):
        return True

    if user.has_permissao(Permissao.ARTIGO_EDITAR_INSTITUICAO.value):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.instituicao_id == user_est.instituicao_id:
            return True

    if user.has_permissao(Permissao.ARTIGO_EDITAR_ESTABELECIMENTO.value):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.estabelecimento_id == user_est.id:
            return True

    if user.has_permissao(Permissao.ARTIGO_EDITAR_SETOR.value):
        if article.setor_id == user.setor_id:
            return True
        if user.extra_setores.filter_by(id=article.setor_id).count():
            return True

    if user.has_permissao(Permissao.ARTIGO_EDITAR_CELULA.value):
        if article.celula_id == user.celula_id:
            return True
        if user.extra_celulas.filter_by(id=article.celula_id).count():
            return True

    return False


def user_can_approve_article(user, article):
    """Verifica se o usuário pode aprovar determinado artigo."""
    try:
        from .models import Article  # type: ignore  # pragma: no cover
    except ImportError:  # pragma: no cover - fallback for direct execution
        from models import Article

    if not isinstance(article, Article):
        return False

    if user.has_permissao("admin") or user.has_permissao(Permissao.ARTIGO_APROVAR_TODAS.value):
        return True

    if user.has_permissao(Permissao.ARTIGO_APROVAR_INSTITUICAO.value):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.instituicao_id == user_est.instituicao_id:
            return True

    if user.has_permissao(Permissao.ARTIGO_APROVAR_ESTABELECIMENTO.value):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.estabelecimento_id == user_est.id:
            return True

    if user.has_permissao(Permissao.ARTIGO_APROVAR_SETOR.value):
        if article.setor_id == user.setor_id:
            return True
        if user.extra_setores.filter_by(id=article.setor_id).count():
            return True

    if user.has_permissao(Permissao.ARTIGO_APROVAR_CELULA.value):
        if article.celula_id == user.celula_id:
            return True
        if user.extra_celulas.filter_by(id=article.celula_id).count():
            return True

    return False


def user_can_review_article(user, article):
    """Verifica se o usuário pode revisar determinado artigo."""
    try:
        from .models import Article  # type: ignore  # pragma: no cover
    except ImportError:  # pragma: no cover - fallback for direct execution
        from models import Article

    if not isinstance(article, Article):
        return False

    if user.has_permissao("admin") or user.has_permissao(Permissao.ARTIGO_REVISAR_TODAS.value):
        return True

    if user.has_permissao(Permissao.ARTIGO_REVISAR_INSTITUICAO.value):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.instituicao_id == user_est.instituicao_id:
            return True

    if user.has_permissao(Permissao.ARTIGO_REVISAR_ESTABELECIMENTO.value):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.estabelecimento_id == user_est.id:
            return True

    if user.has_permissao(Permissao.ARTIGO_REVISAR_SETOR.value):
        if article.setor_id == user.setor_id:
            return True
        if user.extra_setores.filter_by(id=article.setor_id).count():
            return True

    if user.has_permissao(Permissao.ARTIGO_REVISAR_CELULA.value):
        if article.celula_id == user.celula_id:
            return True
        if user.extra_celulas.filter_by(id=article.celula_id).count():
            return True

    return False

def user_can_view_article(user, article):
    """Verifica se o usuário tem permissão para visualizar o artigo."""
    try:
        from .models import Article  # type: ignore  # pragma: no cover
    except ImportError:  # pragma: no cover - fallback for direct execution
        from models import Article
    try:
        from .enums import ArticleVisibility
    except ImportError:  # pragma: no cover - fallback for direct execution
        from enums import ArticleVisibility

    if not isinstance(article, Article):
        return False

    if user_can_edit_article(user, article):
        return True

    if user.has_permissao("admin"):
        return True

    if (
        user.has_permissao(Permissao.ARTIGO_APROVAR_TODAS.value)
        or user.has_permissao(Permissao.ARTIGO_REVISAR_TODAS.value)
    ):
        return True

    if (
        user.has_permissao(Permissao.ARTIGO_APROVAR_INSTITUICAO.value)
        or user.has_permissao(Permissao.ARTIGO_REVISAR_INSTITUICAO.value)
    ):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.instituicao_id == user_est.instituicao_id:
            return True

    if (
        user.has_permissao(Permissao.ARTIGO_APROVAR_ESTABELECIMENTO.value)
        or user.has_permissao(Permissao.ARTIGO_REVISAR_ESTABELECIMENTO.value)
    ):
        user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
        if user_est and article.estabelecimento_id == user_est.id:
            return True

    if (
        user.has_permissao(Permissao.ARTIGO_APROVAR_SETOR.value)
        or user.has_permissao(Permissao.ARTIGO_REVISAR_SETOR.value)
    ):
        if article.setor_id == user.setor_id or user.extra_setores.filter_by(id=article.setor_id).count():
            return True

    if (
        user.has_permissao(Permissao.ARTIGO_APROVAR_CELULA.value)
        or user.has_permissao(Permissao.ARTIGO_REVISAR_CELULA.value)
    ):
        if article.celula_id == user.celula_id or user.extra_celulas.filter_by(id=article.celula_id).count():
            return True

    # Extras concedidos especificamente
    if article.extra_users.filter_by(id=user.id).count():
        return True
    if user.celula_id and article.extra_celulas.filter_by(id=user.celula_id).count():
        return True

    vis = article.visibility

    user_est = user.estabelecimento or (user.celula.estabelecimento if user.celula else None)
    user_setor = user.setor or (user.celula.setor if user.celula else None)

    if vis is ArticleVisibility.INSTITUICAO:
        if user_est and article.instituicao_id == user_est.instituicao_id:
            return True
    elif vis is ArticleVisibility.ESTABELECIMENTO:
        if user_est and article.estabelecimento_id == user_est.id:
            return True
    elif vis is ArticleVisibility.SETOR:
        if user_setor and article.setor_id == user_setor.id:
            return True
    elif vis is ArticleVisibility.CELULA:
        # Usa a célula explícita de visibilidade se definida; caso contrário,
        # faz fallback para a célula do autor (para artigos antigos que não
        # possuam o campo vis_celula_id preenchido)
        cel_id = article.vis_celula_id or article.celula_id
        if user.celula_id == cel_id:
            return True
        # Considera usuários atribuídos a várias células
        if cel_id and user.extra_celulas.filter_by(id=cel_id).count():
            return True

    return False


def eligible_review_notification_users(article):
    """Retorna os usuários que devem ser notificados sobre a revisão do artigo."""
    try:
        from .models import User  # type: ignore  # pragma: no cover
    except ImportError:  # pragma: no cover - fallback for direct execution
        from models import User

    return [
        u for u in User.query.all()
        if user_can_approve_article(u, article) or user_can_review_article(u, article)
    ]
